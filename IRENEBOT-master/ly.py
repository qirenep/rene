import os
path = "./ly"
file_list = os.listdir(path)
print(file_list)
from openpyxl import load_workbook
results = []
for file_name_raw in file_list:
    file_name = "./ly/" + file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active
    result = []
    result.append(file_name_raw)
    result.append(ws['C2'].value)
    result.append(ws['C3'].value)
    result.append(ws['E2'].value)
    result.append(ws['E3'].value)
    results.append(result)
print(results)
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
for i in results:
    ws.append(i)
wb.save("results.xlsx")
