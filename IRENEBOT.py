import asyncio
import discord
import requests
import urllib
import bs4
import re
import random
import math
import sys, operator
import pandas as pd
import urllib.request
import urllib.parse
import youtube_dl
import functools
import itertools
from classes.converters import Platform, Battletag
from async_timeout import timeout
from discord.ext import commands
from openpyxl import load_workbook
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
from utilsmy.http import Fetch, PlayerNotFound
from utilsmy.embed import Embeds, NoCompetitiveStats
token = "YOUR-TOKEN"

class Main(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
    async def battle_else(self, ctx: commands.Context, duoRecord1, embed, duoCenter1):
        duoRat1 = duoRecord1.find("span", {"class": "value"})
        duoRat = duoRat1.text.strip()  # ----레이팅----
        duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
        duoRank = duoRank1.text.strip()  # ----등급----
        print(duoRank)
        embed.add_field(name='레이팅', value=duoRat, inline=False)
        embed.add_field(name='등급', value=duoRank, inline=False)


        duoStat = duoCenter1.find("div", {"class": "stats"})

        duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
        duoKD2 = duoKD1.find("p", {"class": "value"})
        duoKD = duoKD2.text.strip()  # ----킬뎃----
        duoKdSky1 = duoStat.find("span", {"class": "top"})
        duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
        print(duoKD)
        print(duoKdSky)
        embed.add_field(name='킬뎃,킬뎃상위', value=duoKD+" "+duoKdSky, inline=False)

        duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
        duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
        duoWinRat = duoWinRat2.text.strip()  # ----승률----
        duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
        duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
        print(duoWinRat)
        print(duoWinRatSky)
        embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

        duoHead1 = duoStat.find("div", {"class": "headshots"})
        duoHead2 = duoHead1.find("p", {"class": "value"})
        duoHead = duoHead2.text.strip()  # ----헤드샷----
        duoHeadSky1 = duoHead1.find("span", {"class": "top"})
        duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
        print(duoHead)
        print(duoHeadSky)
        embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
        await ctx.send(embed=embed)

    @commands.command(name='안녕')
    async def _hi(self, ctx: commands.Context):
        await ctx.send("hi")
        
    @commands.command(name='롤')
    async def _lol(self, ctx: commands.Context, *, location: str):
        print(location)
        enc_location = urllib.parse.quote(location)

        url = "http://www.op.gg/summoner/userName=" + enc_location
        html = urllib.request.urlopen(url)

        bsObj = bs4.BeautifulSoup(html, "html.parser")
        rank1 = bsObj.find("div", {"class": "TierRankInfo"})
        rank2 = rank1.find("div", {"class": "TierRank"})
        rank3 = rank2.text# 티어표시 (브론즈1,2,3,4,5 등등)
        rank4 = ''.join(rank3.split())
        print(rank4)
        if rank4 != "Unranked":
            jumsu1 = rank1.find("div", {"class": "TierInfo"})
            jumsu2 = jumsu1.find("span", {"class": "LeaguePoints"})
            jumsu3 = jumsu2.text
            jumsu4 = jumsu3.strip()#점수표시 (11LP등등)
            print(jumsu4)

            winlose1 = jumsu1.find("span", {"class": "WinLose"})
            winlose2 = winlose1.find("span", {"class": "wins"})
            winlose2_1 = winlose1.find("span", {"class": "losses"})
            winlose2_2 = winlose1.find("span", {"class": "winratio"})

            winlose2txt = winlose2.text
            winlose2_1txt = winlose2_1.text
            winlose2_2txt = winlose2_2.text #승,패,승률 나타냄  200W 150L Win Ratio 55% 등등

            print(winlose2txt + " " + winlose2_1txt + " " + winlose2_2txt)
        embed = discord.Embed(
            title='롤 정보',
            description='롤 정보입니다.',
            colour=discord.Colour.green()
        )
        if rank4 == "Unranked":
            embed.add_field(name='당신의 티어', value=rank3, inline=False)
            embed.add_field(name='-당신은 언랭-', value="언랭은 더이상의 정보를 제공하지 않습니다.", inline=False)
            await ctx.send(embed=embed)
        else:
         embed.add_field(name='당신의 티어', value=rank3, inline=False)
         embed.add_field(name='당신의 LP(점수)', value=jumsu3, inline=False)
         embed.add_field(name='당신의 승,패 정보', value=winlose2txt+" "+winlose2_1txt, inline=False)
         embed.add_field(name='당신의 승률', value=winlose2_2txt, inline=False)
         await ctx.send(embed=embed)

    @commands.command(name='프사')
    async def avatar(self, ctx: commands.Context, member: discord.Member):
        """Returns members avatar."""
        author = ctx.message.author.mention
        mention = member.mention
    	
        avatar = "{0} here is {1}'s avatar"

        u = member.avatar_url
        embed = discord.Embed(description=avatar.format(author, mention), colour=discord.Colour.blue())
        embed.set_image(u)
        await ctx.send(embed=embed)

    @commands.command(name='배그솔로')
    async def _battlesolo(self, ctx: commands.Context, *, location: str):
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/"+enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("div", {"class": "overview"})
        duoRecord1 = duoCenter1.text
        duoRecord = duoRecord1.strip()
        embed = discord.Embed(
            title='배그솔로 정보',
            description='배그솔로 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == "No record":
            print("솔로 경기가 없습니다.")
            embed.add_field(name='배그를 한판이라도 해주세요', value='솔로 경기 전적이 없습니다..', inline=False)
            await ctx.send(embed=embed)

        else:
            await self.battle_else(ctx=ctx, duoRecord1=duoRecord1, embed=embed, duoCenter1=duoCenter1)
    @commands.command(name='배그듀오')
    async def _battleduo(self, ctx: commands.Context, *, location: str):
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "duo modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        embed = discord.Embed(
            title='배그듀오 정보',
            description='배그듀오 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('듀오 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='듀오 경기 전적이 없습니다..', inline=False)
            await ctx.send(embed=embed)

        else:
            await self.battle_else(ctx=ctx, duoRecord1=duoRecord1, embed=embed, duoCenter1=duoCenter1)
            
    @commands.command(name='배그스쿼드')
    async def _battlefour(self, ctx: commands.Context, *, location: str):
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "squad modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        embed = discord.Embed(
            title='배그스쿼드 정보',
            description='배그스쿼드 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('스쿼드 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='스쿼드 경기 전적이 없습니다..', inline=False)
            await ctx.send(embed=embed)

        else:
            await self.battle_else(ctx=ctx, duoRecord1=duoRecord1, embed=embed, duoCenter1=duoCenter1)
        
    @commands.command(name='급식')
    async def _food(self, ctx: commands.Context, *, location1: str):
        learn = location1.split(" ")
        loca = learn[0]
        location2 = learn[1]
        lc = ' '
        print(loca)
        print(location2)
        wb = load_workbook('code.xlsx')
        ws = wb.active
        ro = ws['B14:C30']
        for i in ws.rows :
            if(operator.eq(i[0].value, location2)):
                if(operator.eq(i[1].value, loca)):
                    schoolcode1 = i[3].value
        for i in ro:
            if(operator.eq(i[0].value, loca)):
                lc = i[1].value
        meal_notice = (
                        "```css\n"
                        "[-] 2018년 1월 2일 인 경우 18012 로 보낼 것.\n"
                        "[-] 2018년 10월 1일 인 경우 18101 로 보낼 것.\n"
                        "```"
                        )
                
        request = meal_notice + '\n' + '날짜를 보내주세요...'
        request_e = discord.Embed(title="Send to Me", description=request, color=0xcceeff)
        print(schoolcode1)
        await ctx.send(embed=request_e)
        def get_diet(code, ymd, weekday):
                schMmealScCode = code #int 1조식2중식3석식
                schYmd = ymd #str 요청할 날짜 yyyy.mm.dd
                if weekday == 5 or weekday == 6: #토요일,일요일 버림
                    element = " " #공백 반환
                else:
                    num = weekday + 1 #int 요청할 날짜의 요일 0월1화2수3목4금5토6일 파싱한 데이터의 배열이 일요일부터 시작되므로 1을 더해줍니다.
                    URL = "http://stu." + lc + ".go.kr/sts_sci_md01_001.do?" + "schulCode=" + schoolcode1 + "&schulCrseScCode=4" + "&schulKndScCode=04" + "&schMmealScCode=%d&schYmd=%s" % (schMmealScCode, schYmd)
    
                    #http://stu.AAA.go.kr/ 관할 교육청 주소 확인해주세요.
                    #schulCode= 학교고유코드
                    #schulCrseScCode= 1유치원2초등학교3중학교4고등학교
                    #schulKndScCode= 01유치원02초등학교03중학교04고등학교

                    #기존 get_html 함수부분을 옮겨왔습니다.
                    html = ""
                    resp = requests.get(URL)
                    if resp.status_code == 200 : #사이트가 정상적으로 응답할 경우
                        html = resp.text
                    soup = BeautifulSoup(html, 'html.parser')
                    element_data = soup.find_all("tr")
                    element_data = element_data[2].find_all('td')
                    try:
                        element = str(element_data[num])

                        #filter
                        element_filter = ['[', ']', '<td class="textC last">', '<td class="textC">', '</td>', '&amp;', '(h)', '.']
                        for element_string in element_filter :
                            element = element.replace(element_string, '')
                        #줄 바꿈 처리
                        element = element.replace('<br/>', '\n')
                        #모든 공백 삭제
                        element = re.sub(r"\d", "", element)

                    #급식이 없을 경우
                    except:
                        element = " " # 공백 반환
                return element

        async def print_get_meal(local_date, local_weekday, message):
            l_diet = get_diet(2, local_date, local_weekday)
            d_diet = get_diet(3, local_date, local_weekday)

            if len(l_diet) == 1:
                embed = discord.Embed(title="No Meal", description="급식이 없습니다.", color=0x00ff00)
                await ctx.send(embed=embed)
            elif len(d_diet) == 1:
                lunch = local_date + " 중식\n" + l_diet
                embed = discord.Embed(title="Lunch", description=lunch, color=0x00ff00)
                await ctx.send(embed=embed)
            else:
                lunch = local_date + " 중식\n" + l_diet
                dinner = local_date + " 석식\n" + d_diet
                embed= discord.Embed(title="Lunch", description=lunch, color=0x00ff00)
                await ctx.send(embed=embed)
                embed = discord.Embed(title="Dinner", description=dinner, color=0x00ff00)
                await ctx.send(embed=embed)

                    
        def pred(m):
            return m.author == ctx.author and m.channel == ctx.channel
        try:
            meal_date = await bot.wait_for('message', check=pred, timeout=15.0)
                #입력이 없을 경우
        except asyncio.TimeoutError:
            longtimemsg = discord.Embed(title="In 15sec", description='15초내로 입력해주세요. 다시시도 : $g', color=0xff0000)
            await message.channel.send(embed=longtimemsg)
            return
        else:
            meal_date = str(meal_date.content) # 171121
            meal_date = '20' + meal_date[:2] + '.' + meal_date[2:4] + '.' + meal_date[4:6] # 2017.11.21

            s = meal_date.replace('.', ', ') # 2017, 11, 21

                #한자리수 달인 경우를 해결하기위함
            if int(s[6:8]) < 10:
                s = s.replace(s[6:8], s[7:8])

            ss = "datetime.datetime(" + s + ").weekday()"
            try:
                whatday = eval(ss)
            except:
                warnning = discord.Embed(title="Plz Retry", description='올바른 값으로 다시 시도하세요 : $g', color=0xff0000)
                await ctx.send(embed=warnning)
                return

            await print_get_meal(meal_date, whatday, ctx)
        if __name__ == '__pred__':
            pred(m)
    @commands.command(name='영화순위')
    async def _movie(self, ctx: commands.Context):
        # http://ticket2.movie.daum.net/movie/movieranklist.aspx
        i1 = 0 # 랭킹 string값
        embed = discord.Embed(
            title = "영화순위",
            description = "영화순위입니다.",
            colour= discord.Color.red()
        )
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'http://ticket2.movie.daum.net/movie/movieranklist.aspx'
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        moviechartBase = bsObj.find('div', {'class': 'main_detail'})
        moviechart1 = moviechartBase.find('ul', {'class': 'list_boxthumb'})
        moviechart2 = moviechart1.find_all('li')

        for i in range(0, 10):
            i1 = i1+1
            stri1 = str(i1) # i1은 영화랭킹을 나타내는데 사용됩니다
            print()
            print(i)
            print()
            moviechartLi1 = moviechart2[i]  # ------------------------- 1등랭킹 영화---------------------------
            moviechartLi1Div = moviechartLi1.find('div', {'class': 'desc_boxthumb'})  # 영화박스 나타내는 Div
            moviechartLi1MovieName1 = moviechartLi1Div.find('strong', {'class': 'tit_join'})
            moviechartLi1MovieName = moviechartLi1MovieName1.text.strip()  # 영화 제목
            print(moviechartLi1MovieName)

            moviechartLi1Ratting1 = moviechartLi1Div.find('div', {'class': 'raking_grade'})
            moviechartLi1Ratting2 = moviechartLi1Ratting1.find('em', {'class': 'emph_grade'})
            moviechartLi1Ratting = moviechartLi1Ratting2.text.strip()  # 영화 평점
            print(moviechartLi1Ratting)

            moviechartLi1openDay1 = moviechartLi1Div.find('dl', {'class': 'list_state'})
            moviechartLi1openDay2 = moviechartLi1openDay1.find_all('dd')  # 개봉날짜, 예매율 두개포함한 dd임
            moviechartLi1openDay3 = moviechartLi1openDay2[0]
            moviechartLi1Yerating1 = moviechartLi1openDay2[1]
            moviechartLi1openDay = moviechartLi1openDay3.text.strip()  # 개봉날짜
            print(moviechartLi1openDay)
            moviechartLi1Yerating = moviechartLi1Yerating1.text.strip()  # 예매율 ,랭킹변동
            print(moviechartLi1Yerating)  # ------------------------- 1등랭킹 영화---------------------------
            print()
            embed.add_field(name='---------------랭킹'+stri1+'위---------------', value='\n영화제목 : '+moviechartLi1MovieName+'\n영화평점 : '+moviechartLi1Ratting+'점'+'\n개봉날짜 : '+moviechartLi1openDay+'\n예매율,랭킹변동 : '+moviechartLi1Yerating, inline=False) # 영화랭킹


        await ctx.send(embed=embed)
        
        
    @commands.command(name='등록')
    async def _lyjoin(self, ctx: commands.Context):
        learn = message.content.split(maxsplit=1)
        artist_name = learn[1]

        # track information sheet
        artists_list = []
        titles_list = [] # song's title
        trackId_list = []
        number_of_lines_list = [] # lyric's number of lines
        song_list = []

        # 가사 분석을 순수 한글로만 이루어진 가사 데이터로만 진행하고 싶을때 사용하기 위함입니다.
        has_alphabet_list = [] # check if lyric has alphabet

        # lyric data sheet
        lyrics_list = []

        # lyric's sentences sheet
        line_list = [] # index for sentence's line
        sentence_data_list = [] # list for sentence data 
        lyrics_sentence_trackId_list = [] # list for sentence's track identifier

        # pattern that I wanna replace
        rep = {'<br>':'\n', '<br/>':'\n', 'amp;':''}#, '-':' ', '_':'', '—':''}
        # use these three lines to do the replacement
        rep = dict((re.escape(k), v) for k, v in rep.items())
        pattern = re.compile("|".join(rep.keys()))
        def remove_tags(lyrics):
            # replace pattern that I defined above
            lyrics = pattern.sub(lambda m: rep[re.escape(m.group(0))], lyrics)
            # remove [text] and <text>    
            lyrics = re.sub("[\<\[].*?[\>\]]", "", lyrics)
            return lyrics

        url1 = "https://music.naver.com/search/search.nhn?query="

        # using parse module -> convert korean to unicode
        url2 = urllib.parse.quote_plus(artist_name)

        full_url = url1 + url2

        html = urllib.request.urlopen(full_url)
         
        artistSoup = BeautifulSoup(html, "lxml")

        # 아티스트 id 추출
        artist = artistSoup.find('dd', class_='desc').find('a')['href']
        index = artist.find('=')
        artist_id = artist[index+1:]

        html = urlopen("https://music.naver.com/artist/track.nhn?artistId=%s&filteringOptions=RELEASE&sorting=newRelease"%artist_id)

        # 1페이지가 최대인경우 a태그가 존재하지 않으므로 수동으로 설정해야함
        max_page = len(BeautifulSoup(html, "lxml").find('div', class_='paginate2').find_all('a'))
        if(max_page==0):
            max_page=1

        for page in range(1, max_page+1, 1):
            html = urlopen("https://music.naver.com/artist/track.nhn?artistId=%s&filteringOptions=RELEASE&sorting=newRelease&page=%d"%(artist_id,page))

            tr_list = BeautifulSoup(html, "lxml").find_all('tr', class_='_tracklist_move')

            for tr in tr_list:
                trackinfo = tr['trackdata'].split('|')
                # tr.find('td', class_='tb_artist').find('span', class_='tit') != None 는 가수 두명 이상의 곡을 제외하고 싶을때 추가하면 됨
                # 가사 포함된 곡만 가져오기, Inst 들어간 제목 거르기, Ver 들어간 제목 거르기, cover 들어간 제목 거르기, MR 들어간 제목 거르기,Remix 거르기,Edit 거르기, 아카펠라 & 어쿠스틱 거르기, 노래 중복 거르기
                # song.find('Feat') == -1 and song.find('With') == -1 이 조건도 포함하면 다른 아티스트가 피쳐링한 곡도 거르고 찾는 아티스트가 다른 아티스트에게 피쳐링한 곡도 거름
                # DJ, By 포함된 곡 거르고 싶으면 song.find('DJ') == -1 and song.find('By') == -1
                song = tr.find('td', class_='tb_name').find('span',class_='tit').text.strip('\n')
                if(trackinfo[9]=='true' and song.find('Inst') == -1 and song.find('Ver') == -1 and song.find('cover') == -1 and song.find('MR') == -1 and song.find('remix') == -1 and song.find('Remix') == -1 and song.find('REMIX') == -1 and song.find('Remastered') == -1 and song.find('Acappella') == -1 and song.find('Acoustic') == -1 and song.find('Edit') == -1 and song not in song_list):
                    trackId_list.append(trackinfo[0])
                    song_list.append(song)

        for t in trackId_list:
            html = urlopen("https://music.naver.com/lyric/index.nhn?trackId=%s"%t)
         
            trackSoup = BeautifulSoup(html, "lxml")

            title = trackSoup.find('span', class_='ico_play').text
            
            artist = str(trackSoup.find('span', class_='artist').find('a'))

            artist = str(re.findall(r'\>(.*?)\<', artist))[2:-2]

            lyrics = str(trackSoup.find('div', class_='show_lyrics'))

            lyrics = remove_tags(lyrics)

            # split with \n
            lyrics_sentence_list = lyrics.split('\n')

            '''
            # remove punctuation
            lyrics_sentence_list = [re.sub(r'[^\w\s]', '', lyrics, re.UNICODE) for lyrics in lyrics_sentence_list]
            '''

            '''
            # remove row in lyric has alphabet. use this code if you wanna pure korean lyrics
            # 영어로 된 문장 혹은 알파벳이 포함된 문장 전체를 제거합니다. 순수한글로만 이루어진 가사만 얻고 싶을때 사용합시다.
            lyrics_sentence_list = list(filter(lambda w: not re.match(r'[a-zA-Z]+', w), lyrics_sentence_list)) 
            '''

            '''
            # remove only english from row of lyric consist of kor and eng.
            # 한영 혼용 문장에서 오직 영어만을 제거합니다. (한글 부분은 남습니다.) 
            lyrics_sentence_list = [re.sub(r'[a-zA-Z]', '', lyrics) for lyrics in lyrics_sentence_list]
            '''

            # remove blank 
            lyrics_sentence_list = list(filter(lambda a : a != '', lyrics_sentence_list))   
            lyrics_sentence_list = list(filter(lambda a : a.isspace() != True, lyrics_sentence_list)) 
            lyrics_sentence_list = [re.sub(' +', ' ', ly) for ly in lyrics_sentence_list]
            lyrics_sentence_list = [ly.strip() for ly in lyrics_sentence_list]
                
            # remove blank from entire lyric
            lyrics = '\n'.join(lyrics_sentence_list)

            if(lyrics == 'None' or lyrics == '가사가 존재하지 않습니다'):
                lyrics = ''
                lyrics_sentence_list = ['']

            # TrackInfo Sheet
            artists_list.append(artist)
            titles_list.append(title)
            number_of_lines_list.append(len(lyrics_sentence_list))

            # entire lyric
            lyrics_list.append(lyrics)

            # lyric's sentence      
            if t not in lyrics_sentence_trackId_list:     
                line = 1                           
                for ly in lyrics_sentence_list:
                    lyrics_sentence_trackId_list.append(t)
                    sentence_data_list.append(ly)
                    line_list.append(line)
                    line +=1

            check_has_alphabet = "False"
            for ly in lyrics_sentence_list:
                alpahbet_val = re.search('[a-zA-Z]+',ly)
                if(alpahbet_val == None): 
                    pass
                else:
                    if(alpahbet_val[0].isalpha()):
                        check_has_alphabet = "True"
                        break

            has_alphabet_list.append(check_has_alphabet)

        trackInfoSheet_column_list = {
        'Artist': artists_list,
        'Title': titles_list,
        'Track_Id' : trackId_list,   
        'Number_Of_Lines' : number_of_lines_list,
        'Has_English' : has_alphabet_list,
        }

        SentenceSheet_column_list = {
        'Track_Id' : lyrics_sentence_trackId_list,
        'Lyric_Sentence' : sentence_data_list,
        'Line' : line_list,
        }

        df_track_info = pd.DataFrame.from_dict(trackInfoSheet_column_list, orient='index')
        df_track_info = df_track_info.transpose()

        df_sentence_data = pd.DataFrame.from_dict(SentenceSheet_column_list, orient='index')
        df_sentence_data = df_sentence_data.transpose()

        writer = pd.ExcelWriter("%s.xlsx" % artist_name)

        df_track_info.to_excel(writer, sheet_name='track_info', startrow=1, header=True)
        df_sentence_data.to_excel(writer, sheet_name='Sentence_data', startrow=1, header=True)

        writer.save()

    @commands.command(name='검색')
    async def _lysearch(self, ctx: commands.Context):
        learn = message.content.split(maxsplit=1)
        base = learn[1]
        
        wb = load_workbook('ly.xlsx')
        track_sheet = wb.worksheets[0]
        sentence_sheet = wb.worksheets[1]
        max_row = sentence_sheet.max_row

        sentence_list = []
        track_id_list = []

        for i in range(2, max_row+1):
            sentence_list.append(sentence_sheet['C%d'%i].value)
            track_id_list.append(sentence_sheet['B%d'%i].value)    

        max_row = track_sheet.max_row

        track_song_info_dict = {}
        track_artist_info_dict = {}
        track_num_line_info_dict = {}

        for i in range(2, max_row+1):
            track_song_info_dict[track_sheet['D%d'%i].value] = track_sheet['C%d'%i].value
            track_artist_info_dict[track_sheet['D%d'%i].value] = track_sheet['B%d'%i].value
            track_num_line_info_dict[track_sheet['D%d'%i].value] = track_sheet['E%d'%i].value

        # n-gram 유사도 비교
        def ngram(s, num):
            res = []
            slen = len(s) - num + 1
            for i in range(slen):
                ss = s[i:i+num]
                res.append(ss)
            return res
        def diff_ngram(sa, sb, num):
            a = ngram(sa, num)
            b = ngram(sb, num)
            r = []
            cnt = 0
            for i in a:
                for j in b:
                    if i == j:
                        cnt += 1
                        r.append(i)
            if len(a) > 0:
                return cnt / len(a), r
            else:
                return 0, r

        three_gram_score_list = []
        three_gram_word_list = []

        for s in sentence_list:
            # 3-gram
            r3, word3  = diff_ngram(base, s, 3)
            three_gram_score_list.append(r3)
            three_gram_word_list.append(word3)

        print('\n알쏭달쏭, 하지만 찾고 싶은 노래의 문장!: %s\n'%base)
        await message.author.send('\n알쏭달쏭, 하지만 찾고 싶은 노래의 문장!: %s\n'%base)

        # 3-gram
        three_max_index = three_gram_score_list.index(max(three_gram_score_list)) 
        three_max_track_id = track_id_list[three_max_index]

        tmp_sentence_list = []
        tmp_track_list = []

        await message.author.send('\n분석 결과 가장 유사한 문장: %s'%sentence_list[three_max_index])
        await message.author.send('분석 결과 가장 유사한 곡은 %s의 %s'%(track_artist_info_dict[three_max_track_id], track_song_info_dict[three_max_track_id]))
        await message.author.send('해당 곡의 전체 가사를 보여드릴게요\n\n')
        print('\n분석 결과 가장 유사한 문장: %s'%sentence_list[three_max_index])
        print('분석 결과 가장 유사한 곡은 %s의 %s'%(track_artist_info_dict[three_max_track_id], track_song_info_dict[three_max_track_id]))

        max_song_line_number = track_num_line_info_dict[three_max_track_id]
        max_song_start_index = track_id_list.index(three_max_track_id)

        max_song_entire = ''
        for i in range(0, max_song_line_number):
            max_song_entire += sentence_list[max_song_start_index + i] + '\n'
            
        await message.author.send(max_song_entire)

        tmp_sentence_list.append(sentence_list[three_max_index])
        tmp_track_list.append(three_max_track_id)

        # total = 이중리스트
        # total = [[아티스트, 곡, 유사도, 문장, 트랙아이디]]
        total = []

        try:
            for i in range(0, len(three_gram_score_list)):
                if three_gram_score_list[i] > 0.25:
                    if(sentence_list[i] not in tmp_sentence_list):
                        tmp_sentence_list.append(sentence_list[i])
                        tmp_track_list.append(track_id_list[i])
                        total.append([track_artist_info_dict[track_id_list[i]], track_song_info_dict[track_id_list[i]], three_gram_score_list[i], sentence_list[i], track_id_list[i]])

            total.sort(key=lambda x:x[2], reverse=True)

            tmp_track_check_list = [three_max_track_id]

            result = []

            for t in total:
                if t[4] not in tmp_track_check_list:
                    result.append(t)
                    tmp_track_check_list.append(t[4])

            if len(result) > 0:
                await message.author.send('\n유사한 다른 곡들도 확인합니다.\n')
                print('\n유사한 다른 곡들도 확인합니다.\n')
                for r in result:
                    await message.author.send('%s의 %s / 유사한 문장: %s' % (r[0], r[1], r[3]))
        except KeyError:
            sys.exit(1)

    

class Statistics(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.command(aliases=["rating"])
    @commands.guild_only()
    @commands.cooldown(1, 3, commands.BucketType.user)
    async def rank(self, ctx, platform: Platform, *, name: Battletag):
        """
        Returns player rank.
        Platform must be: pc, psn or xbl.
        Name must be a battletag if paltform is pc else type your console online id.
        E.g. -rank pc battletag (pc)
        E.g. -rank psn name (psn or xbl)
        Note: name and battletag are case sensitive.
        """
        async with ctx.typing():
            try:
                fetch = Fetch(platform, name)
                data = await fetch.data()
                fmt = Embeds(data, platform, name)
                if data["private"]:
                    embed = fmt.is_private(ctx)
                else:
                    embed = fmt.rank()
                return await ctx.send(embed=embed)
            except PlayerNotFound:
                await ctx.send("Account not found. Make sure you typed in the correct name.")
            except Exception as ex:
                embed = fmt.exception(ex)
                await ctx.send(embed=embed)

    @commands.command(aliases=["medals"])
    @commands.guild_only()
    @commands.cooldown(1, 3, commands.BucketType.user)
    async def awards(self, ctx, platform: Platform, *, name: Battletag):
        """Returns player awards."""
        async with ctx.typing():
            try:
                fetch = Fetch(platform, name)
                data = await fetch.data()
                fmt = Embeds(data, platform, name)
                if data["private"]:
                    embed = fmt.is_private(ctx)
                else:
                    embed = fmt.awards()
                return await ctx.send(embed=embed)
            except PlayerNotFound:
                await ctx.send("Account not found. Make sure you typed in the correct name.")
            except Exception as ex:
                embed = fmt.exception(ex)
                await ctx.send(embed=embed)

    @commands.command(aliases=["quick"])
    @commands.guild_only()
    @commands.cooldown(1, 3, commands.BucketType.user)
    async def quickplay(self, ctx, platform: Platform, *, name: Battletag):
        """Returns player quickplay stats."""
        async with ctx.typing():
            try:
                fetch = Fetch(platform, name)
                data = await fetch.data()
                fmt = Embeds(data, platform, name)
                if data["private"]:
                    embed = fmt.is_private(ctx)
                else:
                    embed = fmt._stats(ctx)
                return await ctx.send(embed=embed)
            except PlayerNotFound:
                await ctx.send("Account not found. Make sure you typed in the correct name.")
            except Exception as ex:
                embed = fmt.exception(ex)
                await ctx.send(embed=embed)

    @commands.command(aliases=["comp"])
    @commands.guild_only()
    @commands.cooldown(1, 3, commands.BucketType.user)
    async def competitive(self, ctx, platform: Platform, *, name: Battletag):
        """Returns player competitive stats."""
        async with ctx.typing():
            try:
                fetch = Fetch(platform, name)
                data = await fetch.data()
                fmt = Embeds(data, platform, name)
                if data["private"]:
                    embed = fmt.is_private(ctx)
                else:
                    embed = fmt._stats(ctx)
                return await ctx.send(embed=embed)
            except PlayerNotFound:
                await ctx.send("Account not found. Make sure you typed in the correct name.")
            except NoCompetitiveStats:
                await ctx.send("This profile has no competitive stats")
            except Exception as ex:
                embed = fmt.exception(ex)
                await ctx.send(embed=embed)


youtube_dl.utils.bug_reports_message = lambda: ''


class VoiceError(Exception):
    pass


class YTDLError(Exception):
    pass


class YTDLSource(discord.PCMVolumeTransformer):
    YTDL_OPTIONS = {
        'format': 'bestaudio/best',
        'extractaudio': True,
        'audioformat': 'mp3',
        'outtmpl': '%(extractor)s-%(id)s-%(title)s.%(ext)s',
        'restrictfilenames': True,
        'noplaylist': True,
        'nocheckcertificate': True,
        'ignoreerrors': False,
        'logtostderr': False,
        'quiet': True,
        'no_warnings': True,
        'default_search': 'auto',
        'source_address': '0.0.0.0',
    }

    FFMPEG_OPTIONS = {
        'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5',
        'options': '-vn',
    }

    ytdl = youtube_dl.YoutubeDL(YTDL_OPTIONS)

    def __init__(self, ctx: commands.Context, source: discord.FFmpegPCMAudio, *, data: dict, volume: float = 0.5):
        super().__init__(source, volume)

        self.requester = ctx.author
        self.channel = ctx.channel
        self.data = data

        self.uploader = data.get('uploader')
        self.uploader_url = data.get('uploader_url')
        date = data.get('upload_date')
        self.upload_date = date[6:8] + '.' + date[4:6] + '.' + date[0:4]
        self.title = data.get('title')
        self.thumbnail = data.get('thumbnail')
        self.description = data.get('description')
        self.duration = self.parse_duration(int(data.get('duration')))
        self.tags = data.get('tags')
        self.url = data.get('webpage_url')
        self.views = data.get('view_count')
        self.likes = data.get('like_count')
        self.dislikes = data.get('dislike_count')
        self.stream_url = data.get('url')

    def __str__(self):
        return '**{0.title}** by **{0.uploader}**'.format(self)

    @classmethod
    async def create_source(cls, ctx: commands.Context, search: str, *, loop: asyncio.BaseEventLoop = None):
        loop = loop or asyncio.get_event_loop()

        partial = functools.partial(cls.ytdl.extract_info, search, download=False, process=False)
        data = await loop.run_in_executor(None, partial)

        if data is None:
            raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        if 'entries' not in data:
            process_info = data
        else:
            process_info = None
            for entry in data['entries']:
                if entry:
                    process_info = entry
                    break

            if process_info is None:
                raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        webpage_url = process_info['webpage_url']
        partial = functools.partial(cls.ytdl.extract_info, webpage_url, download=False)
        processed_info = await loop.run_in_executor(None, partial)

        if processed_info is None:
            raise YTDLError('Couldn\'t fetch `{}`'.format(webpage_url))

        if 'entries' not in processed_info:
            info = processed_info
        else:
            info = None
            while info is None:
                try:
                    info = processed_info['entries'].pop(0)
                except IndexError:
                    raise YTDLError('Couldn\'t retrieve any matches for `{}`'.format(webpage_url))

        return cls(ctx, discord.FFmpegPCMAudio(info['url'], **cls.FFMPEG_OPTIONS), data=info)

    @staticmethod
    def parse_duration(duration: int):
        minutes, seconds = divmod(duration, 60)
        hours, minutes = divmod(minutes, 60)
        days, hours = divmod(hours, 24)

        duration = []
        if days > 0:
            duration.append('{} days'.format(days))
        if hours > 0:
            duration.append('{} hours'.format(hours))
        if minutes > 0:
            duration.append('{} minutes'.format(minutes))
        if seconds > 0:
            duration.append('{} seconds'.format(seconds))

        return ', '.join(duration)


class Song:
    __slots__ = ('source', 'requester')

    def __init__(self, source: YTDLSource):
        self.source = source
        self.requester = source.requester

    def create_embed(self):
        embed = (discord.Embed(title='Now playing',
                               description='```css\n{0.source.title}\n```'.format(self),
                               color=discord.Color.blurple())
                 .add_field(name='Duration', value=self.source.duration)
                 .add_field(name='Requested by', value=self.requester.mention)
                 .add_field(name='Uploader', value='[{0.source.uploader}]({0.source.uploader_url})'.format(self))
                 .add_field(name='URL', value='[Click]({0.source.url})'.format(self))
                 .set_thumbnail(url=self.source.thumbnail))

        return embed


class SongQueue(asyncio.Queue):
    def __getitem__(self, item):
        if isinstance(item, slice):
            return list(itertools.islice(self._queue, item.start, item.stop, item.step))
        else:
            return self._queue[item]

    def __iter__(self):
        return self._queue.__iter__()

    def __len__(self):
        return self.qsize()

    def clear(self):
        self._queue.clear()

    def shuffle(self):
        random.shuffle(self._queue)

    def remove(self, index: int):
        del self._queue[index]


class VoiceState:
    def __init__(self, bot: commands.Bot, ctx: commands.Context):
        self.bot = bot
        self._ctx = ctx

        self.current = None
        self.voice = None
        self.next = asyncio.Event()
        self.songs = SongQueue()

        self._loop = False
        self._volume = 0.5
        self.skip_votes = set()

        self.audio_player = bot.loop.create_task(self.audio_player_task())

    def __del__(self):
        self.audio_player.cancel()

    @property
    def loop(self):
        return self._loop

    @loop.setter
    def loop(self, value: bool):
        self._loop = value

    @property
    def volume(self):
        return self._volume

    @volume.setter
    def volume(self, value: float):
        self._volume = value

    @property
    def is_playing(self):
        return self.voice and self.current

    async def audio_player_task(self):
        while True:
            self.next.clear()

            if not self.loop:
                # Try to get the next song within 3 minutes.
                # If no song will be added to the queue in time,
                # the player will disconnect due to performance
                # reasons.
                try:
                    async with timeout(180):  # 3 minutes
                        self.current = await self.songs.get()
                except asyncio.TimeoutError:
                    self.bot.loop.create_task(self.stop())
                    return

            self.current.source.volume = self._volume
            self.voice.play(self.current.source, after=self.play_next_song)
            await self.current.source.channel.send(embed=self.current.create_embed())

            await self.next.wait()

    def play_next_song(self, error=None):
        if error:
            raise VoiceError(str(error))

        self.next.set()

    def skip(self):
        self.skip_votes.clear()

        if self.is_playing:
            self.voice.stop()

    async def stop(self):
        self.songs.clear()

        if self.voice:
            await self.voice.disconnect()
            self.voice = None


class Music(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.voice_states = {}

    def get_voice_state(self, ctx: commands.Context):
        state = self.voice_states.get(ctx.guild.id)
        if not state:
            state = VoiceState(self.bot, ctx)
            self.voice_states[ctx.guild.id] = state

        return state

    def cog_unload(self):
        for state in self.voice_states.values():
            self.bot.loop.create_task(state.stop())

    def cog_check(self, ctx: commands.Context):
        if not ctx.guild:
            raise commands.NoPrivateMessage('This command can\'t be used in DM channels.')

        return True

    async def cog_before_invoke(self, ctx: commands.Context):
        ctx.voice_state = self.get_voice_state(ctx)

    async def cog_command_error(self, ctx: commands.Context, error: commands.CommandError):
        await ctx.send('An error occurred: {}'.format(str(error)))

    @commands.command(name='join', invoke_without_subcommand=True)
    async def _join(self, ctx: commands.Context):
        """Joins a voice channel."""

        destination = ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='summon')
    @commands.has_permissions(manage_guild=True)
    async def _summon(self, ctx: commands.Context, *, channel: discord.VoiceChannel = None):
        """Summons the bot to a voice channel.
        If no channel was specified, it joins your channel.
        """

        if not channel and not ctx.author.voice:
            raise VoiceError('You are neither connected to a voice channel nor specified a channel to join.')

        destination = channel or ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='leave', aliases=['disconnect'])
    @commands.has_permissions(manage_guild=False)
    async def _leave(self, ctx: commands.Context):
        """Clears the queue and leaves the voice channel."""

        if not ctx.voice_state.voice:
            return await ctx.send('Not connected to any voice channel.')

        await ctx.voice_state.stop()
        del self.voice_states[ctx.guild.id]

    @commands.command(name='volume')
    async def _volume(self, ctx: commands.Context, *, volume: int):
        """Sets the volume of the player."""

        if not ctx.voice_state.is_playing:
            return await ctx.send('Nothing being played at the moment.')

        if 0 > volume > 100:
            return await ctx.send('Volume must be between 0 and 100')

        ctx.voice_state.volume = volume / 100
        await ctx.send('Volume of the player set to {}%'.format(volume))

    @commands.command(name='now', aliases=['current', 'playing'])
    async def _now(self, ctx: commands.Context):
        """Displays the currently playing song."""

        await ctx.send(embed=ctx.voice_state.current.create_embed())

    @commands.command(name='pause')
    @commands.has_permissions(manage_guild=False)
    async def _pause(self, ctx: commands.Context):
        """Pauses the currently playing song."""

        if not ctx.voice_state.is_playing and ctx.voice_state.voice.is_playing():
            ctx.voice_state.voice.pause()
            await ctx.message.add_reaction('⏯')

    @commands.command(name='resume')
    @commands.has_permissions(manage_guild=False)
    async def _resume(self, ctx: commands.Context):
        """Resumes a currently paused song."""

        if not ctx.voice_state.is_playing and ctx.voice_state.voice.is_paused():
            ctx.voice_state.voice.resume()
            await ctx.message.add_reaction('⏯')

    @commands.command(name='stop')
    @commands.has_permissions(manage_guild=False)
    async def _stop(self, ctx: commands.Context):
        """Stops playing song and clears the queue."""

        ctx.voice_state.songs.clear()

        if not ctx.voice_state.is_playing:
            ctx.voice_state.voice.stop()
            await ctx.message.add_reaction('⏹')

    @commands.command(name='skip')
    async def _skip(self, ctx: commands.Context):
        """Vote to skip a song. The requester can automatically skip.
        3 skip votes are needed for the song to be skipped.
        """

        if not ctx.voice_state.is_playing:
            return await ctx.send('Not playing any music right now...')

        voter = ctx.message.author
        if voter == ctx.voice_state.current.requester:
            await ctx.message.add_reaction('⏭')
            ctx.voice_state.skip()

        elif voter.id not in ctx.voice_state.skip_votes:
            ctx.voice_state.skip_votes.add(voter.id)
            total_votes = len(ctx.voice_state.skip_votes)

            if total_votes >= 3:
                await ctx.message.add_reaction('⏭')
                ctx.voice_state.skip()
            else:
                await ctx.send('Skip vote added, currently at **{}/3**'.format(total_votes))

        else:
            await ctx.send('You have already voted to skip this song.')

    @commands.command(name='queue', aliases=['q'])
    async def _queue(self, ctx: commands.Context, *, page: int = 1):
        """Shows the player's queue.
        You can optionally specify the page to show. Each page contains 10 elements.
        """

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Empty queue.')

        items_per_page = 10
        pages = math.ceil(len(ctx.voice_state.songs) / items_per_page)

        start = (page - 1) * items_per_page
        end = start + items_per_page

        queue = ''
        for i, song in enumerate(ctx.voice_state.songs[start:end], start=start):
            queue += '`{0}.` [**{1.source.title}**]({1.source.url})\n'.format(i + 1, song)

        embed = (discord.Embed(description='**{} tracks:**\n\n{}'.format(len(ctx.voice_state.songs), queue))
                 .set_footer(text='Viewing page {}/{}'.format(page, pages)))
        await ctx.send(embed=embed)
        

    @commands.command(name='shuffle')
    async def _shuffle(self, ctx: commands.Context):
        """Shuffles the queue."""

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Empty queue.')

        ctx.voice_state.songs.shuffle()
        await ctx.message.add_reaction('✅')

    @commands.command(name='remove')
    async def _remove(self, ctx: commands.Context, index: int):
        """Removes a song from the queue at a given index."""

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Empty queue.')

        ctx.voice_state.songs.remove(index - 1)
        await ctx.message.add_reaction('✅')

    @commands.command(name='loop')
    async def _loop(self, ctx: commands.Context):
        """Loops the currently playing song.
        Invoke this command again to unloop the song.
        """

        if not ctx.voice_state.is_playing:
            return await ctx.send('Nothing being played at the moment.')

        # Inverse boolean value to loop and unloop.
        ctx.voice_state.loop = not ctx.voice_state.loop
        await ctx.message.add_reaction('✅')

    @commands.command(name='play', aliases=['p'])
    async def _play(self, ctx: commands.Context, *, search: str):
        """Plays a song.
        If there are songs in the queue, this will be queued until the
        other songs finished playing.
        This command automatically searches from various sites if no URL is provided.
        A list of these sites can be found here: https://rg3.github.io/youtube-dl/supportedsites.html
        """

        if not ctx.voice_state.voice:
            await ctx.invoke(self._join)

        async with ctx.typing():
            try:
                source = await YTDLSource.create_source(ctx, search, loop=self.bot.loop)
            except YTDLError as e:
                await ctx.send('An error occurred while processing this request: {}'.format(str(e)))
            else:
                song = Song(source)

                await ctx.voice_state.songs.put(song)
                await ctx.send('Enqueued {}'.format(str(source)))

    @_join.before_invoke
    @_play.before_invoke
    async def ensure_voice_state(self, ctx: commands.Context):
        if not ctx.author.voice or not ctx.author.voice.channel:
            raise commands.CommandError('You are not connected to any voice channel.')

        if ctx.voice_client:
            if ctx.voice_client.channel != ctx.author.voice.channel:
                raise commands.CommandError('Bot is already in a voice channel.')


bot = commands.Bot('r' and 'R')
bot.add_cog(Music(bot))
bot.add_cog(Main(bot))
bot.add_cog(Statistics(bot))

@bot.event
async def on_ready():
    print('Logged in as:\n{0.user.name}\n{0.user.id}'.format(bot))
    await bot.change_presence(activity=discord.Game(name="연애", type=0))
bot.run('KEY')
